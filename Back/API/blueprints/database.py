from flask import Flask, request, jsonify, render_template
from flask import Blueprint, request, jsonify, send_file
from datetime import datetime
from .dao import cur
from psycopg2 import sql, DatabaseError
import re
import re
import io
import csv
import zipfile
import json
import openpyxl
from flask import request, jsonify, send_file
from psycopg2 import sql, DatabaseError

database_bp = Blueprint('database', __name__)

@database_bp.route('/tables', methods=['GET'])
def list_tables():
    """
    GET /api/greenlake-eval/tables
    Devuelve todas las tablas (BASE TABLES) del esquema público y otros esquemas de usuario,
    incluyendo su descripción si existe, y el conteo de columnas.
    """
    try:
        # Query to get table names, descriptions, and column counts
        cur.execute("""
            SELECT
                nspname AS schema,
                relname AS table,
                obj_description(pg_class.oid) AS description,
                (SELECT count(*) FROM information_schema.columns WHERE table_schema = 'public' AND table_name = relname) AS column_count
            FROM pg_class
            JOIN pg_namespace ON pg_namespace.oid = pg_class.relnamespace
            WHERE relkind = 'r' -- Only ordinary tables
              AND nspname = 'public' -- Only public schema
            ORDER BY nspname, relname;
        """)
        rows = cur.fetchall()
    except Exception as e:
        return jsonify({
            "metadata": {
                "status": "error",
                "timestamp": datetime.now().isoformat() + "Z"
            },
            "error": f"Error al consultar la base de datos: {e}"
        }), 500

    # Format each row into a dictionary
    tables = [
        {"schema": schema, "table": table, "description": description, "column_count": column_count}
        for schema, table, description, column_count in rows
    ]

    return jsonify({
        "metadata": {
            "status": "success",
            "timestamp": datetime.now().isoformat() + "Z"
        },
        "results": tables
    })

@database_bp.route('/columns/<table_name>', methods=['GET'])
def get_columns(table_name):
    """
    GET /api/greenlake-eval/columns/<table_name>
    Devuelve todas las columnas y sus tipos de datos de una tabla específica en el esquema público.
    """
    try:
        cur.execute("""
            SELECT column_name, data_type
              FROM information_schema.columns
             WHERE table_schema = 'public' AND table_name = %s
             ORDER BY ordinal_position;
        """, (table_name,))
        rows = cur.fetchall()

    except Exception as e:
        return jsonify({
            "metadata": {
                "status": "error",
                "timestamp": datetime.now().isoformat() + "Z"
            },
            "error": f"Error al consultar la base de datos: {e}"
        }), 500

    if not rows:
        return jsonify({
            "metadata": {
                "status": "error",
                "timestamp": datetime.now().isoformat() + "Z"
            },
            "error": f"La tabla '{table_name}' no existe o no tiene columnas."
        }), 404

    # Formateamos en una lista de diccionarios {name, type}
    columns = [{"column_name": name, "data_type": dtype} for name, dtype in rows]

    return jsonify({
        "metadata": {
            "status": "success",
            "timestamp": datetime.now().isoformat() + "Z"
        },
        "results": columns
    })

@database_bp.route('/export', methods=['POST'])
def export_tables():
    # 1) Validar payload
    payload = request.get_json(force=True, silent=True)
    if not payload or 'tables' not in payload:
        return jsonify({
            "status": "error",
            "message": "Debe enviar un JSON con la clave 'tables'."
        }), 400
    
    tables = payload['tables']
    export_type = payload.get('type', 'csv')  # Default to CSV if not specified
    
    if export_type not in ['csv', 'json', 'xlsx']:
        return jsonify({
            "status": "error",
            "message": "El tipo de exportación debe ser 'csv', 'json' o 'xlsx'."
        }), 400
    
    if not isinstance(tables, list) or not all(isinstance(t, str) for t in tables):
        return jsonify({
            "status": "error",
            "message": "La clave 'tables' debe ser una lista de strings."
        }), 400
    
    # 2) Validar nombres de tabla (evitar inyección)
    valid_name = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')
    for tbl in tables:
        if not valid_name.match(tbl):
            return jsonify({
                "status": "error",
                "message": f"Nombre de tabla inválido: '{tbl}'."
            }), 400
    
    # 3) Recolectar datos de cada tabla
    table_data = {}
    try:
        for tbl in tables:
            # Construir consulta segura
            query = sql.SQL("SELECT * FROM {}").format(sql.Identifier(tbl))
            cur.execute(query)
            cols = [col.name for col in cur.description]
            rows = cur.fetchall()
            
            # Almacenar datos en formato adecuado según el tipo de exportación
            if export_type == 'csv':
                sio = io.StringIO()
                writer = csv.writer(sio)
                writer.writerow(cols)
                writer.writerows(rows)
                table_data[tbl] = sio.getvalue()
            
            elif export_type == 'json':
                # Convertir filas a lista de diccionarios para JSON
                json_rows = []
                for row in rows:
                    json_rows.append(dict(zip(cols, row)))
                table_data[tbl] = json.dumps(json_rows, default=str)
            
            elif export_type == 'xlsx':
                # Para XLSX guardamos los encabezados y filas para procesarlos después
                table_data[tbl] = {'headers': cols, 'rows': rows}
    
    except DatabaseError as e:
        msg = str(e).split('\n')[0]
        return jsonify({
            "status": "error",
            "message": f"Error al exportar tabla '{tbl}': {msg}"
        }), 400
    
    # 4) Si solo hay una tabla, enviar archivo directamente
    if len(table_data) == 1:
        tbl, content = next(iter(table_data.items()))
        
        if export_type == 'csv':
            mem = io.BytesIO(content.encode('utf-8'))
            mem.seek(0)
            return send_file(
                mem,
                as_attachment=True,
                download_name=f"{tbl}.csv",
                mimetype="text/csv"
            )
        
        elif export_type == 'json':
            mem = io.BytesIO(content.encode('utf-8'))
            mem.seek(0)
            return send_file(
                mem,
                as_attachment=True,
                download_name=f"{tbl}.json",
                mimetype="application/json"
            )
        
        elif export_type == 'xlsx':
            # Crear archivo Excel en memoria
            xlsx_io = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = tbl
            
            # Escribir encabezados
            for col_idx, header in enumerate(content['headers'], 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Escribir datos
            for row_idx, row in enumerate(content['rows'], 2):
                for col_idx, cell_value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            wb.save(xlsx_io)
            xlsx_io.seek(0)
            
            return send_file(
                xlsx_io,
                as_attachment=True,
                download_name=f"{tbl}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    # 5) Si hay varias tablas, empaquetar en ZIP
    mem_zip = io.BytesIO()
    with zipfile.ZipFile(mem_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
        for tbl, content in table_data.items():
            if export_type == 'csv':
                zf.writestr(f"{tbl}.csv", content)
            
            elif export_type == 'json':
                zf.writestr(f"{tbl}.json", content)
            
            elif export_type == 'xlsx':
                # Crear archivo Excel en memoria para cada tabla
                xlsx_io = io.BytesIO()
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = tbl
                
                # Escribir encabezados
                for col_idx, header in enumerate(content['headers'], 1):
                    ws.cell(row=1, column=col_idx, value=header)
                
                # Escribir datos
                for row_idx, row in enumerate(content['rows'], 2):
                    for col_idx, cell_value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                wb.save(xlsx_io)
                zf.writestr(f"{tbl}.xlsx", xlsx_io.getvalue())
    
    mem_zip.seek(0)
    return send_file(
        mem_zip,
        as_attachment=True,
        download_name=f"export_tables.{export_type}.zip",
        mimetype="application/zip"
    )
    
@database_bp.route('/export_query', methods=['POST'])
def export_query():
    # 1) Validar payload
    payload = request.get_json(force=True, silent=True)
    if not payload or 'query' not in payload:
        return jsonify({
            "status": "error",
            "message": "Debe enviar un JSON con la clave 'query'."
        }), 400

    query = payload['query']
    export_type = payload.get('type', 'csv')  # CSV por defecto

    # 2) Validar tipo de exportación
    if export_type not in ['csv', 'json', 'xlsx']:
        return jsonify({
            "status": "error",
            "message": "El tipo de exportación debe ser 'csv', 'json' o 'xlsx'."
        }), 400

    # 3) Asegurarse de que sea solo SELECT (para evitar modificaciones de BD)
    if not re.match(r'^\s*SELECT\b', query, re.IGNORECASE):
        return jsonify({
            "status": "error",
            "message": "Solo se permiten consultas SELECT."
        }), 400

    # 4) Ejecutar consulta
    try:
        cur.execute(query)
        cols = [col.name for col in cur.description]
        rows = cur.fetchall()
    except DatabaseError as e:
        msg = str(e).split('\n')[0]
        return jsonify({
            "status": "error",
            "message": f"Error al ejecutar la consulta: {msg}"
        }), 400

    # 5) Generar contenido según formato
    if export_type == 'csv':
        sio = io.StringIO()
        writer = csv.writer(sio)
        writer.writerow(cols)
        writer.writerows(rows)
        data_bytes = sio.getvalue().encode('utf-8')
        filename = "query_export.csv"
        mimetype = "text/csv"

    elif export_type == 'json':
        json_rows = [dict(zip(cols, row)) for row in rows]
        content = json.dumps(json_rows, default=str)
        data_bytes = content.encode('utf-8')
        filename = "query_export.json"
        mimetype = "application/json"

    else:  # xlsx
        xlsx_io = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados"

        # Encabezados
        for col_idx, header in enumerate(cols, 1):
            ws.cell(row=1, column=col_idx, value=header)
        # Filas
        for row_idx, row in enumerate(rows, 2):
            for col_idx, cell_value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        wb.save(xlsx_io)
        xlsx_io.seek(0)
        data_bytes = xlsx_io.read()
        filename = "query_export.xlsx"
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # 6) Enviar el archivo
    mem = io.BytesIO(data_bytes)
    mem.seek(0)
    return send_file(
        mem,
        as_attachment=True,
        download_name=filename,
        mimetype=mimetype
    )